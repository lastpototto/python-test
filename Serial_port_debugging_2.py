import time,threading,serial,serial.tools.list_ports,tkinter as tk,numpy as np,crcmod,re,openpyxl,os
from tkinter import ttk
from tkinter import *
from tkinter import messagebox
from ttkbootstrap import Style
from tkinter import scrolledtext
from threading import Timer
from binascii import *
from PIL import Image
from PIL import ImageTk
from openpyxl import Workbook

# global UART          #全局型式保存串口句柄
# global RX_THREAD     #全局型式保存串口接收函数
# global gui           #全局型式保存GUI句柄
# global count1
# global d
# global open_cnt  #记录串口是否打开过

d = ['00.0000000','00.0000000','00.0000000','00.0000000','00.00000000','00.000000000','00.00000000','00.000000000']  #初始化数组
count1 = 0

def ISHEX(data):        #判断输入字符串是否为十六进制
    if len(data)%2:  #整除则为16进制，取余为1则为非16进制
        return False
    for item in data:
        if item not in '0123456789ABCDEFabcdef': #循环判断数字和字符
            return False
    return True

def uart_open_close(fun,com,bund):  #串口打开关闭控制
    global UART
    global RX_THREAD
    global open_cnt
    global lock
    if fun==1:#打开串口
        try:
           UART = serial.Serial(com, 115200, timeout=0.2)  # 提取串口号和波特率并打开串口
           if UART.isOpen(): #判断是否打开成功
               open_cnt = 1
               lock = threading.Lock() 
               RX_THREAD = UART_RX_TREAD('URX1',lock)  #开启数据接收进程
               RX_THREAD.setDaemon(True)               #开启守护进程 主进程结束后接收进程也关闭 会报警告
               RX_THREAD.start()
               RX_THREAD.resume()
               return True
        except:
            return False
        return False
    else:                   #关闭串口
        RX_THREAD.pause()
        UART.close()
        print("关闭串口")

def uart_tx(data,isHex=False):          #串口发送数据
    try:
        if  UART.isOpen():  #发送前判断串口状态 避免错误
            print("uart_send=" + data)
            # gui.tx_rx_cnt(tx=len(data)) #发送计数
            if isHex:   #十六进制发送
                data_bytes = bytes.fromhex(data)
                return UART.write(bytes(data_bytes))
            else:      #字符发送
                return UART.write(data.encode('gb2312'))
    except:#错误返回
        messagebox.showinfo('错误', '发送失败')   #弹窗
        gui.task()

'''GUI'''''''''''''''''''''''''''''''''''''''''''''''''''''''''
class GUI():
    def __init__(self):
        self.root = tk.Tk()
        self.root.iconbitmap('favicon.ico')
        self.root.title('温度显示程序  软件版本v1.5')             #窗口名称
        self.root.resizable()   #窗口拉伸
        self.root.geometry("1100x570+150+150")         #尺寸位置
        self.interface()
        Style(theme='united') #主题修改 可选['cyborg', 'journal', 'darkly', 'flatly' 'solar', 'minty', 'litera', 'united', 'pulse', 'cosmo', 'lumen', 'yeti', 'superhero','sandstone']

    def interface(self):
        """"界面编写位置"""
        #--------------------------------操作区域-----------------------------#
        self.fr1=Frame(self.root)
        self.fr1.place(x=0,y=0,width=300,height=570)     #区域1位置尺寸

        self.fr2=Frame(self.root)          #区域1 容器  relief   groove=凹  ridge=凸
        self.fr2.place(x=300,y=0,width=900,height=570)     #区域1位置尺寸

        self.var_cb1 = StringVar()  
        self.comb1 = ttk.Combobox(self.fr1,textvariable=self.var_cb1,font=("微软雅黑",11,'bold'))
        self.comb1['values'] = list(serial.tools.list_ports.comports()) #列出可用串口
        self.comb1.current(0)  # 设置默认选项个数 0开始
        self.comb1.place(x=10,y=10,width=280,height=34)

        # print('**********可用串口***********')

        self.var_bt1 = StringVar()              #打开串口按钮
        self.var_bt1.set("打开串口")
        self.btn1 = Button(self.fr1,textvariable=self.var_bt1,font=("微软雅黑",12,'bold'),command=self.uart_opn_close) #绑定 uart_opn_close 方法
        self.btn1.place(x=20,y=50,width=250,height=50)

        self.lb21 = tk.Label(self.fr1, text='(请首先打开串口)',font=("宋体",11))   #提示
        self.lb21.place(x=80,y=105,width=120,height=20)

        self.btn3 = Button(self.fr1, text='清空温度显示区',font=("微软雅黑",13,'bold'),command = self.txt_clr) #绑定清空方法
        self.btn3.place(x=20,y=130,width=250,height=50)    #清空按钮

        self.btn41 = Button(self.fr1, text='开启设备',font=("微软雅黑",13,'bold'),command=self.uart_send1) #绑定发送启动指令
        self.btn41.place(x=20,y=190,width=120,height=50)

        self.btn42 = Button(self.fr1, text='关闭设备',font=("微软雅黑",13,'bold'),command=self.uart_send2) #绑定发送启动指令
        self.btn42.place(x=150,y=190,width=120,height=50)

        self.var_bt7 = StringVar()    #监控按钮
        self.var_bt7.set('开启温度监控')
        self.btn7 = Button(self.fr1, textvariable=self.var_bt7,font=("微软雅黑",13,'bold'),command=self.task)   #绑定监控方法
        self.btn7.place(x=20,y=250,width=250,height=50)

        tk.Label(self.fr1,text='设定间隔',font=('微软雅黑',12,'bold')).place(x=20,y=310,width=80,height=30)  #提示
        
        self.e3 = tk.Entry(self.fr1,font=('微软雅黑',12,'bold'))   #输入时间间隔框
        self.e3.place(x=100,y=310,width=58,height=30)
        self.e3.insert(0,"400")

        tk.Label(self.fr1,text='毫秒',font=('微软雅黑',12,'bold')).place(x=160,y=310,width=50,height=30) #提示单位
        
        tk.Label(self.fr1,text='（若需修改间隔时间，请先关闭监控）',font=('微软雅黑',10)).place(x=0,y=340,width=260,height=30)  #提示

        self.lb16 = Label(self.fr1,text='程序版本号:',font=('微软雅黑',12,'bold'))  #版本号
        self.lb16.place(x=10,y=380,width=100,height=40)
        
        self.version = StringVar()   #版本号2212
        self.version_number = Label(self.fr1,textvariable=self.version,font=('微软雅黑',12,'bold'))
        self.version_number.place(x=120,y=380,width=80,height=40)
        
        self.lbs =Label(self.fr1,text='设定温度注意事项：',font=('宋体',12,'bold'))  #提示
        self.lbs.place(x=10,y=430,width=190,height=40)

        self.lbs =Label(self.fr1,text='1.设定温度不能超过小数点后一位',font=('宋体',10,'bold'))  #提示
        self.lbs.place(x=22,y=460,width=220,height=40)

        self.lbs1 =Label(self.fr1,text='2.首先关闭监控，再清空温度显示区',font=('宋体',10,'bold'))  #提示
        self.lbs1.place(x=20,y=490,width=240,height=40)

        self.lbs2 =Label(self.fr1,text='最后输入设定温度，设定即发送',font=('宋体',10,'bold'))  #提示
        self.lbs2.place(x=30,y=520,width=220,height=40)

#  ————————————————————————————————FR2区——————————————————————————————————————————————————————
        
        self.lb5 = Label(self.fr2,text='设定温度:',font=('微软雅黑',12,'bold'))   #设定温度
        self.lb5.place(x=20,y=25,width=70,height=20)

        self.btn2 = Button(self.fr2, text='设定',font=('微软雅黑',12,'bold'),command=self.crc16Add1)
        self.btn2.place(x=220,y=15,width=80,height=40)    #发送设定温度按钮

        self.txt_ex = scrolledtext.ScrolledText(self.fr2,font=('微软雅黑',14,'bold'))    #设定温度
        self.txt_ex.place(relheight=0.06,relwidth=0.11,relx=0.12,rely=0.03) #比例计算控件尺寸和位置 (x与y放大：向右向下)

        self.lb6 = Label(self.fr2,text='低温阈值:',font=('微软雅黑',12,'bold'))
        self.lb6.place(x=20,y=80,width=70,height=20)
        
        self.btn4 = Button(self.fr2, text='设定',font=('微软雅黑',12,'bold'),command=self.crc16Add2)
        self.btn4.place(x=220,y=70,width=80,height=40)    #发送设定温度按钮

        self.txt_e1x = scrolledtext.ScrolledText(self.fr2,font=('微软雅黑',14,'bold'))
        self.txt_e1x.place(relheight=0.06,relwidth=0.11,relx=0.12,rely=0.13)

        self.lb7 = Label(self.fr2,text='高温阈值:',font=('微软雅黑',12,'bold'))
        self.lb7.place(x=20,y=140,width=70,height=20)

        self.btn5 = Button(self.fr2, text='设定',font=('微软雅黑',12,'bold'),command=self.crc16Add3)
        self.btn5.place(x=220,y=130,width=80,height=40)    #发送设定温度按钮

        self.txt_e2x = scrolledtext.ScrolledText(self.fr2,font=('微软雅黑',14,'bold'))
        self.txt_e2x.place(relheight=0.06,relwidth=0.11,relx=0.12,rely=0.235)

        self.lb8 = Label(self.fr2,text='温度校准:',font=('微软雅黑',12,'bold'))
        self.lb8.place(x=20,y=200,width=70,height=20)

        self.btn9 = Button(self.fr2, text='设定',font=('微软雅黑',12,'bold'),command=self.crc16Add4)
        self.btn9.place(x=220,y=190,width=80,height=40)    #发送设定温度按钮

        self.txt_e3x = scrolledtext.ScrolledText(self.fr2,font=('微软雅黑',14,'bold'))
        self.txt_e3x.place(relheight=0.06,relwidth=0.11,relx=0.12,rely=0.34)

        self.lb12 = Label(self.fr2,text='水冷出水口:',font=('微软雅黑',12,'bold'))
        self.lb12.place(x=335,y=25,width=85,height=20)

        self.txt_e7x = scrolledtext.ScrolledText(self.fr2,font=('微软雅黑',14,'bold'))
        self.txt_e7x.place(relheight=0.06,relwidth=0.11,relx=0.48,rely=0.03)

        self.lb10 = Label(self.fr2,text='水冷热端:',font=('微软雅黑',12,'bold'))
        self.lb10.place(x=340,y=80,width=80,height=20)

        self.txt_e5x = scrolledtext.ScrolledText(self.fr2,font=('微软雅黑',14,'bold'))
        self.txt_e5x.place(relheight=0.06,relwidth=0.11,relx=0.48,rely=0.13)
        
        self.lb14 = Label(self.fr2,text='环境温度:',font=('微软雅黑',12,'bold'))
        self.lb14.place(x=340,y=140,width=80,height=20)

        self.txt_e10x = scrolledtext.ScrolledText(self.fr2,font=('微软雅黑',14,'bold'))
        self.txt_e10x.place(relheight=0.06,relwidth=0.11,relx=0.48,rely=0.235)

        self.lb13 = Label(self.fr2,text='拍子出水口:',font=('微软雅黑',12,'bold'))
        self.lb13.place(x=335,y=200,width=85,height=20)
        
        self.txt_e11x = scrolledtext.ScrolledText(self.fr2,font=('微软雅黑',14,'bold'))
        self.txt_e11x.place(relheight=0.06,relwidth=0.11,relx=0.48,rely=0.34)

        self.txt_rx = scrolledtext.ScrolledText(self.fr2,font=('微软雅黑',12))    #接收
        self.txt_rx.place(relheight=0.3,relwidth=0.68,relx=0,rely=0.47) #比例计算控件尺寸和位置

        self.txt_tx = Text(self.fr2,font=('微软雅黑',12,'bold'))   #手动发送
        self.txt_tx.place(relheight=0.06,relwidth=0.48,relx=0,rely=0.83) #比例计算控件尺寸位置

        self.btn6 = Button(self.fr2, text='发送',font=('微软雅黑',12,'bold'),command=self.uart_send)  #绑定发送方法
        self.btn6.place(relheight=0.06,relwidth=0.16,relx=0.5,rely=0.83)   #发送按钮

        self.photo1 = Image.open("error1.png")
        self.photo1 = self.photo1.resize((25,25))
        self.photo3 = ImageTk.PhotoImage(self.photo1)

        self.photo2 = Image.open("勾.png")
        self.photo2 = self.photo2.resize((25,25))
        self.photo4 = ImageTk.PhotoImage(self.photo2)

        self.bj1 = Label(self.fr2, text='制冷开关：',font=('微软雅黑',14,'bold'), anchor='w')  #时钟
        self.bj1.place(relheight=0.05, relwidth=0.12, relx=0.7, rely=0.05)

        self.bj2 = Label(self.fr2, text='风扇报警：',font=('微软雅黑',14,'bold'), anchor='w')  #时钟
        self.bj2.place(relheight=0.05, relwidth=0.12, relx=0.7, rely=0.15)

        self.bj3 = Label(self.fr2, text='水泵报警：',font=('微软雅黑',14,'bold'), anchor='w')  #时钟
        self.bj3.place(relheight=0.05, relwidth=0.12, relx=0.7, rely=0.25)

        self.bj4 = Label(self.fr2, text='探头报警：',font=('微软雅黑',14,'bold'), anchor='w')  #时钟
        self.bj4.place(relheight=0.05, relwidth=0.12, relx=0.7, rely=0.35)

        self.bj5 = Label(self.fr2, text='低温报警：',font=('微软雅黑',14,'bold'), anchor='w')  #时钟
        self.bj5.place(relheight=0.05, relwidth=0.12, relx=0.7, rely=0.45)
        
        self.bj6 = Label(self.fr2, text='高温报警：',font=('微软雅黑',14,'bold'), anchor='w')  #时钟
        self.bj6.place(relheight=0.05, relwidth=0.12, relx=0.7, rely=0.55)

        self.bj7 = Label(self.fr2, text='水流报警：',font=('微软雅黑',14,'bold'), anchor='w')  #时钟
        self.bj7.place(relheight=0.05, relwidth=0.12, relx=0.7, rely=0.65)

        self.bj8 = Label(self.fr2, text='水位报警：',font=('微软雅黑',14,'bold'), anchor='w')  #时钟
        self.bj8.place(relheight=0.05, relwidth=0.12, relx=0.7, rely=0.75)

        self.lb4 = Label(self.fr2, text=' ', anchor='w',relief=GROOVE)  #时钟
        self.lb4.place(relheight=0.05, relwidth=0.06, relx=0.8, rely=0.935)

        self.lb5 = Label(self.fr2, text='北京时间：', anchor='w')  #时钟
        self.lb5.place(relheight=0.05, relwidth=0.07, relx=0.72, rely=0.935)
#------------------------------------------方法-----------------------------------------------
    
    def gettim(self):#获取时间 未用
            self.timestr = time.strftime("%H:%M:%S")  # 获取当前的时间并转化为字符串
            self.lb4.configure(text=self.timestr)  # 重新设置标签文本
            self.txt_rx.after(1000, self.gettim)     # 每隔1s调用函数 gettime 自身获取时间 GUI自带的定时函数

    def txt_clr(self):#清空显示
        self.txt_tx.delete(0.0, 'end')  # 清空文本框
        self.txt_ex.delete(0.0, 'end')  # 清空文本框
        self.txt_e1x.delete(0.0, 'end')  # 清空文本框
        self.txt_e2x.delete(0.0, 'end')  # 清空文本框
        self.txt_e3x.delete(0.0, 'end')  # 清空文本框
        # self.txt_e5x.delete(0.0, 'end')  # 清空文本框
        # self.txt_e7x.delete(0.0, 'end')  # 清空文本框
        # self.txt_e10x.delete(0.0, 'end')  # 清空文本框
        # self.txt_e11x.delete(0.0, 'end')  # 清空文本框   
           
    def uart_opn_close(self):#打开关闭串口
        if(self.var_bt1.get() == '打开串口'):
            if(uart_open_close(1,str(self.comb1.get())[0:5],115200)==True): #传递下拉框选择的参数 COM号+波特率  【0:5】表示只提取COM号字符
                self.var_bt1.set('关闭串口')                             #改变按键内容
                self.txt_rx.insert(0.0, self.comb1.get() + ' 打开成功\r\n')  # 开头插入
                send_data3 = 'AA 03 00 11 00 03 4C 15'  #发送读取版本指令
                uart_tx(send_data3,True)
            else:
                print("串口打开失败")
                messagebox.showinfo('错误','串口打开失败')
        else:
            uart_open_close(0,'COM1', 115200) #使fun=0，即判断为关闭
            self.var_bt1.set('打开串口')

    def uart_send(self): #发送数据
        send_data = self.txt_tx.get(0.0, 'end').strip()
        send_data = send_data.replace(" ", "").replace("\n", "0A").replace("\r", "0D") #替换空格和回车换行
        if(ISHEX(send_data)==False):
            messagebox.showinfo('错误', '请输入十六进制数')
            return
        uart_tx(send_data,True)
    def uart_send1(self):
            send_data1 = 'AA10000000091200010064001E000A00010226020802BC01F40873'
            self.txt_rx.insert(0.0,'AA 10 00 00 00 09 12 00 01 00 64 00 1E 00 0A 00 01 02 26 02 08 02 BC 01 F4 08 73'+'  发送启动    时间：'+self.timestr+'\r\n')
            uart_tx(send_data1,True)

    def uart_send2(self):
        send_data2 = 'AA06000000009011'
        self.txt_rx.insert(0.0,'AA 06 00 00 00 00 90 11'+'  发送停机    时间：'+self.timestr+'\r\n')
        uart_tx(send_data2,True)

    def crc16Add1(self):  
        lastnum=hex(int((float(self.txt_ex.get('1.0','1.end'))+50)*10))
        self.e2 = 'AA 06 00 05 0'+lastnum[-3:-2]+' '+lastnum[-2:]
        self.e2 =self.e2.upper()
        crc16 = crcmod.mkCrcFun(0x18005, rev=True, initCrc=0xFFFF, xorOut=0x0000)
        data = self.e2.replace(" ", "") #消除空格
        readcrcout = hex(crc16(unhexlify(data))).upper()
        str_list = list(readcrcout)
        if len(str_list) == 5:
            str_list.insert(2, '0')  # 位数不足补0，因为一般最少是5个
        crc_data = "".join(str_list) #用""把数组的每一位结合起来  组成新的字符串
        read = self.e2[:] + ' ' +crc_data[4:] + ' ' + crc_data[2:4] #把源代码和crc校验码连接起来
        uart_tx(str(read),True)   
        self.txt_rx.insert(0.0,str(read)+'  发送设定温度    时间：'+self.timestr+'\r\n')
    
    def crc16Add2(self):  
        lastnum=hex(int((float(self.txt_e1x.get('1.0','1.end'))+50)*10))
        self.e4 = 'AA 06 00 06 0'+lastnum[-3:-2]+' '+lastnum[-2:]
        self.e4 =self.e4.upper()
        crc16 = crcmod.mkCrcFun(0x18005, rev=True, initCrc=0xFFFF, xorOut=0x0000)
        data = self.e4.replace(" ", "") #消除空格
        readcrcout = hex(crc16(unhexlify(data))).upper()
        str_list = list(readcrcout)
        if len(str_list) == 5:
            str_list.insert(2, '0')  # 位数不足补0，因为一般最少是5个
        crc_data = "".join(str_list) #用""把数组的每一位结合起来  组成新的字符串
        read = self.e4[:] + ' ' +crc_data[4:] + ' ' + crc_data[2:4] #把源代码和crc校验码连接起来
        uart_tx(str(read),True)   
        self.txt_rx.insert(0.0,str(read)+'  发送低温阈值    时间：'+self.timestr+'\r\n')

    def crc16Add3(self):  
        lastnum=hex(int((float(self.txt_e2x.get('1.0','1.end'))+50)*10))
        self.e5 = 'AA 06 00 07 0'+lastnum[-3:-2]+' '+lastnum[-2:]
        self.e5 =self.e5.upper()
        crc16 = crcmod.mkCrcFun(0x18005, rev=True, initCrc=0xFFFF, xorOut=0x0000)
        data = self.e5.replace(" ", "") #消除空格
        readcrcout = hex(crc16(unhexlify(data))).upper()
        str_list = list(readcrcout)
        if len(str_list) == 5:
            str_list.insert(2, '0')  # 位数不足补0，因为一般最少是5个
        crc_data = "".join(str_list) #用""把数组的每一位结合起来  组成新的字符串
        read = self.e5[:] + ' ' +crc_data[4:] + ' ' + crc_data[2:4] #把源代码和crc校验码连接起来
        uart_tx(str(read),True)   
        self.txt_rx.insert(0.0,str(read)+'  发送高温阈值    时间：'+self.timestr+'\r\n')

    def crc16Add4(self):  
        lastnum=hex(int((float(self.txt_e3x.get('1.0','1.end'))+50)*10))
        self.e6 = 'AA 06 00 08 0'+lastnum[-3:-2]+' '+lastnum[-2:]
        self.e6 =self.e6.upper()
        crc16 = crcmod.mkCrcFun(0x18005, rev=True, initCrc=0xFFFF, xorOut=0x0000)
        data = self.e6.replace(" ", "") #消除空格
        readcrcout = hex(crc16(unhexlify(data))).upper()
        str_list = list(readcrcout)
        if len(str_list) == 5:
            str_list.insert(2, '0')  # 位数不足补0，因为一般最少是5个
        crc_data = "".join(str_list) #用""把数组的每一位结合起来  组成新的字符串
        read = self.e6[:] + ' ' +crc_data[4:] + ' ' + crc_data[2:4] #把源代码和crc校验码连接起来
        uart_tx(str(read),True)   
        self.txt_rx.insert(0.0,str(read)+'  发送温度校准    时间：'+self.timestr+'\r\n')

    def func(self):   # 需要定时执行的函数
        send_data3 = 'AA0300000012DC1C'
        self.txt_rx.insert(0.0,'AA 03 00 00 00 12 DC 1C'+'  发送读取指令    时间：'+self.timestr+'\r\n')
        uart_tx(send_data3,True)
        if not self.var_bt7.get() == '开启温度监控': 
            self.closemoniter()
    
    def closemoniter(self):
        self.timer=Timer((float(self.e3.get())/1000),self.func)
        self.timer.start()
    
    def task(self):
        if(self.var_bt7.get() == '开启温度监控'): 
            self.var_bt7.set('关闭温度监控')
            send_data3 = 'AA0300000012DC1C'
            self.txt_rx.insert(0.0,'AA 03 00 00 00 12 DC 1C'+'  发送读取    时间：'+self.timestr+'\r\n')
            uart_tx(send_data3,True)
            self.closemoniter()
        else:
            self.var_bt7.set('开启温度监控')
            self.timer.cancel()  #关闭定时器
    
class UART_RX_TREAD(threading.Thread):          #数据接收进程 部分重构
    def __init__(self, name, lock):
        threading.Thread.__init__(self)  #__init__和run()被重写
        self.mName = name
        self.mLock = lock
        self.mEvent = threading.Event()
        
    def run(self): #主函数
        global count1
        global d
        print('开启数据接收\r')
        while True:
            self.mEvent.wait()
            lock.acquire()
            if UART.isOpen():
                rx_buf =  UART.readall()
                if len(rx_buf) == 41:
                    text_list = re.findall(".{2}",rx_buf.hex().upper())
                    new_text = " ".join(text_list)
                    gui.txt_rx.insert(0.0,new_text+'  接收数据    时间：'+gui.timestr+'\r\n')
                    datas=''.join(map(lambda UART:('/x' if len(hex(UART))>=4 else '/x0')+hex(UART)[2:],rx_buf))   #将接收到的数据每一个转换成十六进制，中间用/x分割，便于后面分割
                    new_datas=datas[2:].split('/x') # 由于datas变量中的数据前两个是/x，用split切片
                    need=''.join(new_datas) # 将前面的列表中的数据连接起来
                    need_1=list(need)
                    need_1[63]='1'
                    need_1[62]='0'
                    need=''.join(need_1)
                    lastBJ = need[-6:-4].upper()
                    i=27 
                    while i < 74 :    #锁定所需温度区间数据位  
                        b=int((str(need)[i-1:i+1]),16)   #前两位,16为十六进制
                        c=int((str(need)[i+1:i+3]),16)   #后两位
                        temp=(b*16*16+c-500)/10          #温度值
                        i+=4
                        savenumpy(temp)
                        count1 += 1
                    if count1 == 12 :  # 以12个数据为一组循环
                        if lastBJ[0] == 'E' or lastBJ[0] == 'F' or lastBJ[0] == '8' or lastBJ[0] == '9' or lastBJ[0] == 'A' or lastBJ[0] == 'C' or lastBJ[0] == 'B' or lastBJ[0] == 'D' :
                            label1 = tk.Label(gui.fr2, image=gui.photo4)
                            label1.place(relheight=0.05, relwidth=0.07, relx=0.81, rely=0.05) 
                        else:
                            label1 = tk.Label(gui.fr2, image=gui.photo3)
                            label1.place(relheight=0.05, relwidth=0.07, relx=0.81, rely=0.05)
                        if lastBJ[0] == 'E' or lastBJ[0] == 'F' or lastBJ[0] == '4' or lastBJ[0] == '5' or lastBJ[0] == '6' or lastBJ[0] == 'C' or lastBJ[0] == '7' or lastBJ[0] == 'D' :
                            label2 = tk.Label(gui.fr2, image=gui.photo4)
                            label2.place(relheight=0.05, relwidth=0.07, relx=0.81, rely=0.15)  
                        else:
                            label2 = tk.Label(gui.fr2, image=gui.photo3)
                            label2.place(relheight=0.05, relwidth=0.07, relx=0.81, rely=0.15)
                        if lastBJ[0] == 'E' or lastBJ[0] == 'F' or lastBJ[0] == '2' or lastBJ[0] == '3' or lastBJ[0] == '6' or lastBJ[0] == 'A' or lastBJ[0] == '7' or lastBJ[0] == 'B' :
                            label3 = tk.Label(gui.fr2, image=gui.photo4)
                            label3.place(relheight=0.05, relwidth=0.07, relx=0.81, rely=0.25)
                        else:
                            label3 = tk.Label(gui.fr2, image=gui.photo3)
                            label3.place(relheight=0.05, relwidth=0.07, relx=0.81, rely=0.25)
                        if lastBJ[0] == 'E' or lastBJ[0] == '0' or lastBJ[0] == '2' or lastBJ[0] == '4' or lastBJ[0] == '8' or lastBJ[0] == '6' or lastBJ[0] == 'A' or lastBJ[0] == 'C' :
                            label4 = tk.Label(gui.fr2, image=gui.photo4)
                            label4.place(relheight=0.05, relwidth=0.07, relx=0.81, rely=0.35)
                        else:
                            label4 = tk.Label(gui.fr2, image=gui.photo3)
                            label4.place(relheight=0.05, relwidth=0.07, relx=0.81, rely=0.35)
                        if lastBJ[1] == '0' or lastBJ[1] == '1' or lastBJ[1] == '2' or lastBJ[1] == '4' or lastBJ[1] == '3' or lastBJ[1] == '5' or lastBJ[1] == '6' or lastBJ[1] == '7' :
                            label5 = tk.Label(gui.fr2, image=gui.photo4)    			
                            label5.place(relheight=0.05, relwidth=0.07, relx=0.81, rely=0.45)
                        else:
                            label5 = tk.Label(gui.fr2, image=gui.photo3)    			
                            label5.place(relheight=0.05, relwidth=0.07, relx=0.81, rely=0.45)
                        if lastBJ[1] == '0' or lastBJ[1] == '1' or lastBJ[1] == '2' or lastBJ[1] == '8' or lastBJ[1] == '9' or lastBJ[1] == 'A' or lastBJ[1] == 'B' or lastBJ[1] == '3' :
                            label6 = tk.Label(gui.fr2, image=gui.photo4)    			
                            label6.place(relheight=0.05, relwidth=0.07, relx=0.81, rely=0.55)
                        else:
                            label6 = tk.Label(gui.fr2, image=gui.photo3)    			
                            label6.place(relheight=0.05, relwidth=0.07, relx=0.81, rely=0.55)
                        if lastBJ[1] == '0' or lastBJ[1] == '1' or lastBJ[1] == '4' or lastBJ[1] == '8' or lastBJ[1] == '9' or lastBJ[1] == '5' or lastBJ[1] == 'C' or lastBJ[1] == 'D' :
                            label7 = tk.Label(gui.fr2, image=gui.photo4)    			
                            label7.place(relheight=0.05, relwidth=0.07, relx=0.81, rely=0.65)
                        else:
                            label7 = tk.Label(gui.fr2, image=gui.photo3)    			
                            label7.place(relheight=0.05, relwidth=0.07, relx=0.81, rely=0.65)
                        if lastBJ[1] == '0' or lastBJ[1] == '2' or lastBJ[1] == '4' or lastBJ[1] == '8' or lastBJ[1] == '6' or lastBJ[1] == 'A' or lastBJ[1] == 'C' or lastBJ[1] == 'E' :
                            label8 = tk.Label(gui.fr2, image=gui.photo4)    			
                            label8.place(relheight=0.05, relwidth=0.07, relx=0.81, rely=0.75)
                        else:
                            label8 = tk.Label(gui.fr2, image=gui.photo3)    			
                            label8.place(relheight=0.05, relwidth=0.07, relx=0.81, rely=0.75)
                        gui.txt_ex.insert(0.0,d[0]+'\r\n')
                        gui.txt_e1x.insert(0.0,d[1]+'\r\n')
                        gui.txt_e2x.insert(0.0,d[2]+'\r\n')
                        gui.txt_e3x.insert(0.0,d[3]+'\r\n')
                        gui.txt_e7x.insert(0.0,d[4]+'\r\n')
                        gui.txt_e5x.insert(0.0,d[5]+'\r\n')    
                        gui.txt_e10x.insert(0.0,d[6]+'\r\n')
                        gui.txt_e11x.insert(0.0,d[7]+'\r\n')
                        if os.access("test_xlwt.xlsx", os.F_OK):
                            self.xlsxdata = openpyxl.load_workbook('test_xlwt.xlsx')#检查是否存在文件
                        else:
                            self.xlsxdata=openpyxl.Workbook()
                        sheetnames = self.xlsxdata.get_sheet_names()
                        table = self.xlsxdata.get_sheet_by_name(sheetnames[0])
                        table = self.xlsxdata.active
                        nrows = table.max_row # 获得行数
                        # ncolumns = table.max_column # 获得行数
                        for o in range(0,7):
                            table.cell(nrows+1,o+2).value = d[o] #VALUE为方法,table表格的列与行必须从数字1起，对应第一列与第一行
                            table.cell(nrows+1,10).value = time.strftime("%H:%M:%S")
                        self.xlsxdata.save('test_xlwt.xlsx')  #保存数据并命名
                        count1 = 0
                elif len(rx_buf) == 11:
                        text_list = re.findall(".{2}",rx_buf.hex().upper())
                        new_text = " ".join(text_list)
                        datas=''.join(map(lambda UART:('/x' if len(hex(UART))>=4 else '/x0')+hex(UART)[2:],rx_buf))   #将接收到的数据每一个转换成十六进制，中间用/x分割，便于后面分割
                        new_datas=datas[2:].split('/x') # 由于datas变量中的数据前两个是/x，用split切片
                        need=''.join(new_datas) # 将前面的列表中的数据连接起来
                        need=need.upper()
                        self.version1=int(need[10:14],16)
                        self.version2=int(need[14:18],16)
                        self.version1=str(self.version1)
                        self.version2=str(self.version2)
                        gui.version.set(self.version1+self.version2)
            lock.release()
    def pause(self): #暂停
        self.mEvent.clear()

    def resume(self):#恢复
        self.mEvent.set()

def savenumpy(a):  # 保存接收的数据
    if count1 == 0 or count1 == 1 or count1 == 2 or count1 == 3 or count1 == 4 :
        d[count1]=str(a)+'℃'
    elif count1 == 6 :
        d[5]=str(a)+'℃'
    elif count1 == 10 :
        d[6]=str(a)
    elif count1 == 11 :
        d[7]=str(a)
if __name__ == '__main__':    
    gui = GUI()   #全局类调用
    gui.gettim()  #开启时钟
    gui.root.mainloop()

if __name__ == '__main__':
    if open_cnt == 1:
        UART.close()   #结束关闭 避免下次打开错误
    print('End...')   
